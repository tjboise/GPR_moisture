import openpyxl
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.cm as cm
from scipy.signal import hilbert

wb = openpyxl.load_workbook('GPR measurement data in field.xlsx')

CONDITIONS = {
    '2in_sand': {
        'gpr_sheet': '2 in sand gpr data',
        'moist_sheet': '2 in sand moisture data',
        'pipe_depth_m': 0.40,
        'label': '2" pipe, Sand',
        'b_depth_cm': 35,
    },
    '4in_sand': {
        'gpr_sheet': '4in sand gpr data',
        'moist_sheet': '4in sand mosture data',
        'pipe_depth_m': 0.35,
        'label': '4" pipe, Sand',
        'b_depth_cm': 35,
    },
    '4in_clay': {
        'gpr_sheet': '4in clay gpr data',
        'moist_sheet': '4in clay moisture data',
        'pipe_depth_m': 0.30,
        'label': '4" pipe, Clay',
        'b_depth_cm': 30,
    },
}

def load_condition(gpr_sname, moist_sname):
    gpr_ws = wb[gpr_sname]
    moist_ws = wb[moist_sname]
    gpr_rows = list(gpr_ws.iter_rows(values_only=True))
    moist_rows = list(moist_ws.iter_rows(values_only=True))

    time_ns = np.array([r[0] for r in gpr_rows[1:]], dtype=float)
    dt = time_ns[1] - time_ns[0]
    n_traces = gpr_ws.max_column - 1

    traces = np.zeros((len(time_ns), n_traces))
    for col in range(n_traces):
        for row_i, r in enumerate(gpr_rows[1:]):
            v = r[col + 1]
            traces[row_i, col] = v if v is not None else 0

    moisture = {}
    for row_i in range(1, len(moist_rows)):
        label = moist_rows[row_i][0]
        if label is None:
            continue
        vals = []
        for col in range(1, n_traces + 1):
            v = moist_rows[row_i][col] if col < len(moist_rows[row_i]) else None
            vals.append(float(v) if v is not None else np.nan)
        moisture[str(label)] = np.array(vals)

    return time_ns, dt, traces, moisture, n_traces

def find_first_break(trace, threshold_frac=0.05):
    """Find index of first sample exceeding threshold_frac * peak amplitude."""
    peak = np.max(np.abs(trace))
    threshold = threshold_frac * peak
    above = np.where(np.abs(trace) > threshold)[0]
    return above[0] if len(above) > 0 else 0

def align_traces(traces, time_ns):
    """
    Align all traces by their first break.
    Returns aligned traces on a common zero-referenced time axis.
    """
    dt = time_ns[1] - time_ns[0]
    n_samples, n_traces = traces.shape

    # Find first break index for each trace
    fb_indices = np.array([find_first_break(traces[:, i]) for i in range(n_traces)])

    # Reference: use median first break as time zero
    ref_idx = int(np.median(fb_indices))

    # Shift each trace so its first break aligns with ref_idx
    aligned = np.zeros_like(traces)
    for i in range(n_traces):
        shift = fb_indices[i] - ref_idx
        if shift >= 0:
            aligned[shift:, i] = traces[:n_samples - shift, i]
        else:
            aligned[:n_samples + shift, i] = traces[-shift:, i]

    # New time axis centered at t=0 at ref_idx
    t_aligned = (np.arange(n_samples) - ref_idx) * dt

    return aligned, t_aligned, fb_indices, ref_idx

# ── Figure: Aligned A-scans colored by B-layer moisture ───────────────────
fig, axes = plt.subplots(1, 3, figsize=(17, 6))
fig.suptitle('Aligned A-scans — colored by bottom-layer moisture (%)',
             fontsize=13, fontweight='bold')

for ax, (key, cfg) in zip(axes, CONDITIONS.items()):
    t, dt, traces, moisture, n = load_condition(
        cfg['gpr_sheet'], cfg['moist_sheet'])

    aligned, t_al, fb_idx, ref_idx = align_traces(traces, t)

    # B-layer moisture for coloring
    b_key = [k for k in moisture if 'B' in k or 'b' in k][0]
    b_mo = moisture[b_key]

    valid_mask = ~np.isnan(b_mo)
    mo_vals = b_mo[valid_mask]
    mo_min, mo_max = mo_vals.min(), mo_vals.max()

    cmap = cm.get_cmap('coolwarm')
    norm = plt.Normalize(vmin=mo_min, vmax=mo_max)

    for i in range(n):
        if not valid_mask[i]:
            continue
        color = cmap(norm(b_mo[i]))
        # Normalize trace to [-1, 1] for display
        tr = aligned[:, i]
        peak = np.max(np.abs(tr))
        if peak == 0:
            continue
        tr_n = tr / peak
        ax.plot(tr_n, t_al, color=color, alpha=0.5, lw=0.8)

    ax.set_ylim([12, -1])  # time goes downward, show 0–12 ns after alignment
    ax.set_xlim([-1.3, 1.3])
    ax.axhline(y=0, color='k', lw=1, ls='--', alpha=0.5, label='Time zero (aligned)')
    ax.set_xlabel('Normalized amplitude')
    ax.set_ylabel('Time after alignment (ns)')
    ax.set_title(cfg['label'], fontsize=11)
    ax.grid(True, alpha=0.25)

    sm = plt.cm.ScalarMappable(cmap=cmap, norm=norm)
    sm.set_array([])
    cbar = plt.colorbar(sm, ax=ax)
    cbar.set_label(f'B-layer moisture (%)\n({b_key})', fontsize=8)

plt.tight_layout()
plt.savefig('fig_ascan_aligned.png', dpi=150, bbox_inches='tight')
plt.close()
print('Saved fig_ascan_aligned.png')

# ── Figure: show first-break times distribution ────────────────────────────
fig, axes = plt.subplots(1, 3, figsize=(14, 4))
fig.suptitle('First-break time per trace (before alignment)', fontsize=12, fontweight='bold')

for ax, (key, cfg) in zip(axes, CONDITIONS.items()):
    t, dt, traces, moisture, n = load_condition(
        cfg['gpr_sheet'], cfg['moist_sheet'])

    b_key = [k for k in moisture if 'B' in k or 'b' in k][0]
    b_mo = moisture[b_key]

    fb_idx = np.array([find_first_break(traces[:, i]) for i in range(n)])
    fb_times = fb_idx * dt

    valid = ~np.isnan(b_mo)
    sc = ax.scatter(b_mo[valid], fb_times[valid],
                    c=b_mo[valid], cmap='coolwarm', s=50, edgecolors='white', lw=0.5)
    r = np.corrcoef(b_mo[valid], fb_times[valid])[0, 1]
    ax.set_xlabel(f'B-layer moisture (%)', fontsize=9)
    ax.set_ylabel('First-break time (ns)', fontsize=9)
    ax.set_title(f'{cfg["label"]}\nr = {r:.3f}', fontsize=10)
    ax.grid(True, alpha=0.3)
    plt.colorbar(sc, ax=ax, label='Moisture (%)')

plt.tight_layout()
plt.savefig('fig_firstbreak_vs_moisture.png', dpi=150, bbox_inches='tight')
plt.close()
print('Saved fig_firstbreak_vs_moisture.png')

print('Done.')
