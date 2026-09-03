import openpyxl
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from scipy.signal import stft, find_peaks

wb = openpyxl.load_workbook('GPR measurement data in field.xlsx')

CONDITIONS = {
    '2in_sand': {
        'gpr_sheet': '2 in sand gpr data',
        'moist_sheet': '2 in sand moisture data',
        'label': '2" pipe, Sand',
    },
    '4in_sand': {
        'gpr_sheet': '4in sand gpr data',
        'moist_sheet': '4in sand mosture data',
        'label': '4" pipe, Sand',
    },
    '4in_clay': {
        'gpr_sheet': '4in clay gpr data',
        'moist_sheet': '4in clay moisture data',
        'label': '4" pipe, Clay',
    },
}

def load_gpr(gpr_sname):
    ws = wb[gpr_sname]
    rows = list(ws.iter_rows(values_only=True))
    time_ns = np.array([r[0] for r in rows[1:]], dtype=float)
    n = ws.max_column - 1
    traces = np.zeros((len(time_ns), n))
    for col in range(n):
        for ri, r in enumerate(rows[1:]):
            v = r[col + 1]
            traces[ri, col] = v if v is not None else 0
    return time_ns, traces, n

def load_moisture(moist_sname, n_traces):
    ws = wb[moist_sname]
    rows = list(ws.iter_rows(values_only=True))
    moisture = {}
    for ri in range(1, len(rows)):
        label = rows[ri][0]
        if label is None:
            continue
        vals = [float(rows[ri][c]) if c < len(rows[ri]) and rows[ri][c] is not None
                else np.nan for c in range(1, n_traces + 1)]
        moisture[str(label)] = np.array(vals)
    return moisture

def align_trace(trace, threshold_frac=0.05):
    """Return shift index for time-zero correction."""
    peak = np.max(np.abs(trace))
    above = np.where(np.abs(trace) > threshold_frac * peak)[0]
    return above[0] if len(above) > 0 else 0

FS_HZ   = 1e9 / 0.099609   # ~10.04 GHz in Hz
NPERSEG = 56
NOVERLAP = round(NPERSEG * 0.5)
NFFT     = 256

def compute_stft(trace):
    """
    Compute STFT with the project-standard parameters.
    Returns: freqs (Hz), times (s), magnitude (2D array)
    """
    f, t, Zxx = stft(trace, fs=FS_HZ, nperseg=NPERSEG, noverlap=NOVERLAP,
                     nfft=NFFT, window='hann', boundary=None, padded=False)
    return f, t, np.abs(Zxx)

# ── Figure 1: Example STFT images for dry/mid/wet traces ──────────────────
fig, axes = plt.subplots(3, 6, figsize=(20, 10))
fig.suptitle('STFT Spectrograms of GPR A-scans (dry → wet, by B-layer moisture)',
             fontsize=13, fontweight='bold')

for row_i, (key, cfg) in enumerate(CONDITIONS.items()):
    t_ns, traces, n = load_gpr(cfg['gpr_sheet'])
    moisture = load_moisture(cfg['moist_sheet'], n)
    dt = t_ns[1] - t_ns[0]

    b_key = [k for k in moisture if 'B' in k][0]
    b_mo = moisture[b_key]
    valid = np.where(~np.isnan(b_mo))[0]

    # Pick 6 traces spanning dry to wet
    sorted_idx = valid[np.argsort(b_mo[valid])]
    n_pick = 6
    pick_indices = sorted_idx[np.linspace(0, len(sorted_idx)-1, n_pick, dtype=int)]

    for col_i, idx in enumerate(pick_indices):
        ax = axes[row_i, col_i]
        trace = traces[:, idx].copy()

        # Time-zero correction: shift trace
        fb = align_trace(trace)
        ref = int(np.median([align_trace(traces[:, i]) for i in range(n)]))
        shift = fb - ref
        if shift >= 0:
            aligned = np.concatenate([np.zeros(shift), trace[:len(trace)-shift]])
        else:
            aligned = np.concatenate([trace[-shift:], np.zeros(-shift)])

        f_hz, t_s, mag = compute_stft(aligned)

        # Show 0–5 GHz and full time window (seconds)
        f_mask = f_hz <= 5e9
        t_max_s = t_ns[-1] * 1e-9
        t_mask = t_s <= t_max_s

        im = ax.imshow(mag[np.ix_(f_mask, t_mask)],
                       aspect='auto', origin='lower',
                       extent=[t_s[t_mask][0], t_s[t_mask][-1],
                               f_hz[f_mask][0], f_hz[f_mask][-1]],
                       cmap='RdBu_r', interpolation='nearest')

        mo_val = b_mo[idx]
        ax.set_title(f'B-moist={mo_val:.1f}%', fontsize=8)
        ax.set_xlabel('Time (s)', fontsize=7)
        if col_i == 0:
            ax.set_ylabel(f'{cfg["label"]}\nFreq (Hz)', fontsize=7)
        else:
            ax.set_ylabel('Freq (Hz)', fontsize=7)
        plt.colorbar(im, ax=ax, fraction=0.046, pad=0.04)

plt.tight_layout()
plt.savefig('fig_stft_examples.png', dpi=150, bbox_inches='tight')
plt.close()
print('Saved fig_stft_examples.png')

# ── Figure 2: Check dominant frequency of A-scans ──────────────────────────
fig, axes = plt.subplots(1, 3, figsize=(15, 4))
fig.suptitle('Frequency spectrum of A-scans (FFT magnitude)', fontsize=12, fontweight='bold')

for ax, (key, cfg) in zip(axes, CONDITIONS.items()):
    t_ns, traces, n = load_gpr(cfg['gpr_sheet'])
    dt_s = (t_ns[1] - t_ns[0]) * 1e-9
    freqs = np.fft.rfftfreq(len(t_ns), d=dt_s)

    for i in range(0, n, max(1, n//8)):
        spec = np.abs(np.fft.rfft(traces[:, i]))
        ax.plot(freqs / 1e9, spec / spec.max(), alpha=0.4, lw=0.8)

    ax.set_xlim([0, 5])
    ax.set_xlabel('Frequency (GHz)')
    ax.set_ylabel('Normalized magnitude')
    ax.set_title(cfg['label'])
    ax.grid(True, alpha=0.3)

plt.tight_layout()
plt.savefig('fig_frequency_spectrum.png', dpi=150, bbox_inches='tight')
plt.close()
print('Saved fig_frequency_spectrum.png')

print('Done.')
