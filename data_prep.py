"""
Shared data preparation for all GPR moisture prediction experiments.
Loads all 3 conditions, aligns A-scans (time-zero correction),
and saves train/test splits as numpy arrays.
"""
import openpyxl
import numpy as np
from sklearn.model_selection import train_test_split

EXCEL = 'GPR measurement data in field.xlsx'
RANDOM_SEED = 42

CONDITIONS = [
    {
        'name': '2in_sand',
        'gpr_sheet': '2 in sand gpr data',
        'moist_sheet': '2 in sand moisture data',
        'pipe_depth_m': 0.40,
    },
    {
        'name': '4in_sand',
        'gpr_sheet': '4in sand gpr data',
        'moist_sheet': '4in sand mosture data',
        'pipe_depth_m': 0.35,
    },
    {
        'name': '4in_clay',
        'gpr_sheet': '4in clay gpr data',
        'moist_sheet': '4in clay moisture data',
        'pipe_depth_m': 0.30,
    },
]

# Clay depth labels differ from sand — map to same order: S, T, M, B
MOIST_LAYER_ORDER = ['S', 'T', 'M', 'B']


def load_condition(wb, cfg):
    gpr_ws = wb[cfg['gpr_sheet']]
    moist_ws = wb[cfg['moist_sheet']]
    gpr_rows = list(gpr_ws.iter_rows(values_only=True))
    moist_rows = list(moist_ws.iter_rows(values_only=True))

    time_ns = np.array([r[0] for r in gpr_rows[1:]], dtype=float)
    n_traces = gpr_ws.max_column - 1

    traces = np.zeros((len(time_ns), n_traces))
    for col in range(n_traces):
        for ri, r in enumerate(gpr_rows[1:]):
            v = r[col + 1]
            traces[ri, col] = float(v) if v is not None else 0.0

    # Parse moisture rows — skip None-label rows (average row) and header
    moisture_by_label = {}
    for ri in range(1, len(moist_rows)):
        lbl = moist_rows[ri][0]
        if lbl is None:
            continue
        lbl_str = str(lbl).strip()
        vals = []
        for col in range(1, n_traces + 1):
            v = moist_rows[ri][col] if col < len(moist_rows[ri]) else None
            vals.append(float(v) if v is not None else np.nan)
        moisture_by_label[lbl_str] = np.array(vals)

    # Map to canonical order S, T, M, B
    moisture = np.full((4, n_traces), np.nan)
    for layer_i, layer_key in enumerate(MOIST_LAYER_ORDER):
        for lbl, arr in moisture_by_label.items():
            if lbl.startswith(layer_key):
                moisture[layer_i] = arr
                break

    return time_ns, traces, moisture, n_traces


def find_first_break(trace, threshold_frac=0.05):
    peak = np.max(np.abs(trace))
    if peak == 0:
        return 0
    above = np.where(np.abs(trace) > threshold_frac * peak)[0]
    return int(above[0]) if len(above) > 0 else 0


def align_all(traces):
    n_samples, n_traces = traces.shape
    fb = np.array([find_first_break(traces[:, i]) for i in range(n_traces)])
    ref = int(np.median(fb))
    aligned = np.zeros_like(traces)
    for i in range(n_traces):
        shift = fb[i] - ref
        if shift >= 0:
            aligned[shift:, i] = traces[:n_samples - shift, i]
        else:
            aligned[:n_samples + shift, i] = traces[-shift:, i]
    return aligned, ref


def build_dataset():
    wb = openpyxl.load_workbook(EXCEL)
    all_traces = []
    all_moisture = []   # shape (n_samples, 4) — columns: S, T, M, B
    all_condition = []  # integer label per condition

    for cond_i, cfg in enumerate(CONDITIONS):
        time_ns, traces, moisture, n = load_condition(wb, cfg)
        aligned, _ = align_all(traces)

        for i in range(n):
            mo = moisture[:, i]  # [S, T, M, B]
            # Skip samples with any NaN moisture
            if np.any(np.isnan(mo)):
                continue
            all_traces.append(aligned[:, i])
            all_moisture.append(mo)
            all_condition.append(cond_i)

    X = np.array(all_traces, dtype=np.float32)       # (N, 256)
    y = np.array(all_moisture, dtype=np.float32)      # (N, 4)
    cond = np.array(all_condition, dtype=np.int32)    # (N,)

    # Normalize each trace to [-1, 1] by its own peak
    peaks = np.max(np.abs(X), axis=1, keepdims=True)
    peaks[peaks == 0] = 1.0
    X_norm = X / peaks

    # 80/20 split — stratify by condition so all conditions appear in both sets
    idx = np.arange(len(X_norm))
    idx_train, idx_test = train_test_split(
        idx, test_size=0.2, random_state=RANDOM_SEED, stratify=cond)

    print(f'Total samples: {len(X_norm)}')
    print(f'Train: {len(idx_train)}, Test: {len(idx_test)}')
    print(f'Condition counts: {np.bincount(cond)}')
    print(f'Moisture stats (train):')
    for li, lbl in enumerate(MOIST_LAYER_ORDER):
        vals = y[idx_train, li]
        print(f'  {lbl}: mean={vals.mean():.1f}%, std={vals.std():.1f}%, '
              f'range=[{vals.min():.1f}, {vals.max():.1f}]')

    np.save('X_norm.npy', X_norm)
    np.save('y_moisture.npy', y)
    np.save('cond_labels.npy', cond)
    np.save('idx_train.npy', idx_train)
    np.save('idx_test.npy', idx_test)
    np.save('time_ns.npy', np.linspace(0, (255) * 0.099609, 256))

    print('\nSaved: X_norm.npy, y_moisture.npy, cond_labels.npy, idx_train.npy, idx_test.npy')
    return X_norm, y, cond, idx_train, idx_test


if __name__ == '__main__':
    build_dataset()
